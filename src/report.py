"""Build deterministic report data and Excel workbooks from Pipeline results."""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from enum import Enum
from io import BytesIO
import copy
import json
from operator import index
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.pipeline import PipelineResult, PipelineStatus
from src.validator import ValidationResult

INVALID_REPORT_INPUT = "INVALID_REPORT_INPUT"
INVALID_REPORT_DATA = "INVALID_REPORT_DATA"
EXCEL_EXPORT_ERROR = "EXCEL_EXPORT_ERROR"
EXCEL_CELL_TEXT_TOO_LONG = "EXCEL_CELL_TEXT_TOO_LONG"
INCONSISTENT_REPORT_DATA = "INCONSISTENT_REPORT_DATA"
EXCEL_ROW_LIMIT_EXCEEDED = "EXCEL_ROW_LIMIT_EXCEEDED"

SHEET_NAMES: tuple[str, ...] = (
    "Summary",
    "Validation Issues",
    "Metrics",
    "Diagnostics",
)
SUMMARY_COLUMNS: tuple[str, ...] = ("Section", "Item", "Value")
VALIDATION_ISSUE_COLUMNS: tuple[str, ...] = (
    "level",
    "code",
    "row",
    "field",
    "message",
)

COUNT_COLUMNS: frozenset[str] = frozenset(
    {"impressions", "clicks", "orders", "units_sold", "refunds", "inventory"}
)
USD_COLUMNS: frozenset[str] = frozenset(
    {"sales", "ad_spend", "gmv", "aov", "cpc", "cpa"}
)
PERCENTAGE_COLUMNS: frozenset[str] = frozenset({"ctr", "cvr", "refund_rate"})

COUNT_FORMAT = "#,##0"
USD_FORMAT = "$#,##0.00"
PERCENTAGE_FORMAT = "0.00%"
ROAS_FORMAT = '0.00"x"'
DATE_FORMAT = "yyyy-mm-dd"
VALIDATION_FAILED_MESSAGE = "Not generated because validation failed."
EXCEL_MAX_CELL_TEXT_LENGTH = 32_767
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_DATA_ROWS = EXCEL_MAX_ROWS - 1
MAX_EVIDENCE_DEPTH = 20

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_NOTE_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
_MIN_COLUMN_WIDTH = 8
_MAX_COLUMN_WIDTH = 50
_LONG_TEXT_COLUMNS: frozenset[str] = frozenset({"message", "evidence"})
_EXCEL_NATIVE_DATE_MIN = date(1900, 1, 1)
_EXCEL_NATIVE_DATE_MAX = date(9999, 12, 31)
_EXCEL_EXACT_INTEGER_MAX = 2**53 - 1


class ReportError(Exception):
    """A stable failure raised at the Report layer boundary."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        self.message = message
        super().__init__(f"{code}: {message}")


@dataclass(frozen=True)
class ReportData:
    """Presentation copies derived from exactly one PipelineResult."""

    summary: pd.DataFrame
    validation_issues: pd.DataFrame
    metrics: pd.DataFrame | None
    diagnostics: pd.DataFrame | None


def _copy_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Copy a DataFrame, including mutable values stored in object columns."""

    result = dataframe.copy(deep=True)
    for column in result.columns:
        if pd.api.types.is_object_dtype(result[column].dtype):
            result[column] = result[column].map(copy.deepcopy)
    return result


def _validate_pipeline_result(pipeline_result: object) -> PipelineResult:
    if not isinstance(pipeline_result, PipelineResult):
        raise ReportError(
            INVALID_REPORT_INPUT,
            "pipeline_result 必须是 PipelineResult。",
        )
    if not isinstance(pipeline_result.status, PipelineStatus) or not isinstance(
        pipeline_result.validation, ValidationResult
    ):
        raise ReportError(
            INVALID_REPORT_INPUT,
            "PipelineResult 包含无效的 status 或 validation。",
        )
    if pipeline_result.status is PipelineStatus.SUCCESS:
        if not isinstance(pipeline_result.metrics, pd.DataFrame) or not isinstance(
            pipeline_result.diagnostics, pd.DataFrame
        ):
            raise ReportError(
                INVALID_REPORT_INPUT,
                "SUCCESS PipelineResult 必须包含 Metrics 和 Diagnostics DataFrame。",
            )
    elif pipeline_result.metrics is not None or pipeline_result.diagnostics is not None:
        raise ReportError(
            INVALID_REPORT_INPUT,
            "VALIDATION_FAILED PipelineResult 不应包含 Metrics 或 Diagnostics。",
        )
    return pipeline_result


def _validation_issues_dataframe(validation: ValidationResult) -> pd.DataFrame:
    records = [issue.to_dict() for issue in validation.report.issues]
    if not records:
        return pd.DataFrame(
            {
                "level": pd.Series(dtype="object"),
                "code": pd.Series(dtype="object"),
                "row": pd.Series(dtype="Int64"),
                "field": pd.Series(dtype="object"),
                "message": pd.Series(dtype="object"),
            }
        ).loc[:, VALIDATION_ISSUE_COLUMNS]

    result = pd.DataFrame.from_records(records, columns=VALIDATION_ISSUE_COLUMNS)
    result["row"] = pd.Series(result["row"].tolist(), dtype="Int64")
    return result.loc[:, VALIDATION_ISSUE_COLUMNS]


def _append_code_counts(
    rows: list[dict[str, Any]],
    *,
    section: str,
    codes: list[Any],
) -> None:
    for code, count in Counter(codes).items():
        rows.append({"Section": section, "Item": code, "Value": count})


def _summary_dataframe(pipeline_result: PipelineResult) -> pd.DataFrame:
    validation_report = pipeline_result.validation.report
    metrics = pipeline_result.metrics
    diagnostics = pipeline_result.diagnostics
    rows: list[dict[str, Any]] = [
        {"Section": "Pipeline", "Item": "Status", "Value": pipeline_result.status.value},
        {"Section": "Validation", "Item": "Raw Rows", "Value": validation_report.total_rows},
        {"Section": "Validation", "Item": "Valid Rows", "Value": validation_report.valid_rows},
        {"Section": "Validation", "Item": "Excluded Rows", "Value": validation_report.excluded_rows},
        {"Section": "Validation", "Item": "Warning Rows", "Value": validation_report.warning_rows},
        {"Section": "Validation", "Item": "Fatal Issues", "Value": len(validation_report.fatal_errors)},
        {"Section": "Validation", "Item": "Error Issues", "Value": len(validation_report.errors)},
        {"Section": "Validation", "Item": "Warning Issues", "Value": len(validation_report.warnings)},
        {"Section": "Metrics", "Item": "Metrics Groups", "Value": 0 if metrics is None else len(metrics)},
        {"Section": "Diagnostics", "Item": "Diagnostic Issues", "Value": 0 if diagnostics is None else len(diagnostics)},
    ]
    _append_code_counts(
        rows,
        section="Validation Code Counts",
        codes=[issue.code for issue in validation_report.issues],
    )
    if diagnostics is not None and "code" in diagnostics.columns:
        _append_code_counts(
            rows,
            section="Diagnostic Code Counts",
            codes=diagnostics["code"].tolist(),
        )
    return pd.DataFrame.from_records(rows, columns=SUMMARY_COLUMNS)


def build_report_data(pipeline_result: PipelineResult) -> ReportData:
    """Create an Excel-independent presentation model from a PipelineResult."""

    result = _validate_pipeline_result(pipeline_result)
    return ReportData(
        summary=_summary_dataframe(result),
        validation_issues=_validation_issues_dataframe(result.validation),
        metrics=None if result.metrics is None else _copy_dataframe(result.metrics),
        diagnostics=(
            None
            if result.diagnostics is None
            else _copy_dataframe(result.diagnostics)
        ),
    )


def _validate_report_data(report_data: object) -> ReportData:
    if not isinstance(report_data, ReportData):
        raise ReportError(
            INVALID_REPORT_DATA,
            "report_data 必须是 ReportData。",
        )
    if not isinstance(report_data.summary, pd.DataFrame) or not isinstance(
        report_data.validation_issues, pd.DataFrame
    ):
        raise ReportError(
            INVALID_REPORT_DATA,
            "ReportData 的 summary 和 validation_issues 必须是 DataFrame。",
        )
    for name, value in (
        ("metrics", report_data.metrics),
        ("diagnostics", report_data.diagnostics),
    ):
        if value is not None and not isinstance(value, pd.DataFrame):
            raise ReportError(
                INVALID_REPORT_DATA,
                f"ReportData 的 {name} 必须是 DataFrame 或 None。",
            )
    return report_data


def _raise_inconsistent_report_data(message: str) -> None:
    raise ReportError(INCONSISTENT_REPORT_DATA, message)


def _summary_records(summary: pd.DataFrame) -> list[tuple[Any, Any, Any]]:
    required = set(SUMMARY_COLUMNS)
    if not required.issubset(summary.columns):
        missing = [column for column in SUMMARY_COLUMNS if column not in summary.columns]
        _raise_inconsistent_report_data(
            f"Summary 缺少一致性校验字段：{', '.join(missing)}。"
        )
    positions = {column: summary.columns.get_loc(column) for column in SUMMARY_COLUMNS}
    records: list[tuple[Any, Any, Any]] = []
    for row in summary.itertuples(index=False, name=None):
        section = _native_scalar(row[positions["Section"]])
        item = _native_scalar(row[positions["Item"]])
        if not isinstance(section, str) or not isinstance(item, str):
            _raise_inconsistent_report_data(
                "Summary 的 Section 和 Item 必须是非空文本。"
            )
        records.append((section, item, row[positions["Value"]]))
    return records


def _count_value(value: Any, *, label: str) -> int:
    value = _native_scalar(value)
    if isinstance(value, bool):
        _raise_inconsistent_report_data(f"Summary {label} 不是非负整数。")
    try:
        parsed = index(value)
    except TypeError:
        _raise_inconsistent_report_data(f"Summary {label} 不是非负整数。")
    if parsed < 0:
        _raise_inconsistent_report_data(f"Summary {label} 不是非负整数。")
    return parsed


def _summary_count(
    records: list[tuple[Any, Any, Any]],
    *,
    section: str,
    item: str,
) -> int:
    matches = [
        value
        for current_section, current_item, value in records
        if current_section == section and current_item == item
    ]
    if len(matches) != 1:
        _raise_inconsistent_report_data(
            f"Summary 必须且只能包含一条 {section} / {item}。"
        )
    return _count_value(matches[0], label=f"{section} / {item}")


def _detail_code_counts(
    dataframe: pd.DataFrame | None,
    *,
    detail_name: str,
) -> list[tuple[str, int]]:
    if dataframe is None:
        return []
    if "code" not in dataframe.columns:
        _raise_inconsistent_report_data(f"{detail_name} 缺少 code 字段。")
    codes: list[str] = []
    for value in dataframe["code"].tolist():
        code = _native_scalar(value)
        if not isinstance(code, str):
            _raise_inconsistent_report_data(f"{detail_name} 包含无效 code。")
        codes.append(code)
    return list(Counter(codes).items())


def _summary_code_counts(
    records: list[tuple[Any, Any, Any]],
    *,
    section: str,
) -> list[tuple[str, int]]:
    counts: list[tuple[str, int]] = []
    for current_section, item, value in records:
        if current_section != section:
            continue
        if not isinstance(item, str):
            _raise_inconsistent_report_data(f"Summary {section} 包含无效 Code。")
        counts.append(
            (item, _count_value(value, label=f"{section} / {item}"))
        )
    return counts


def _validate_report_consistency(report_data: ReportData) -> None:
    records = _summary_records(report_data.summary)
    validation_issues = report_data.validation_issues
    for column in ("level", "code"):
        if column not in validation_issues.columns:
            _raise_inconsistent_report_data(
                f"Validation Issues 缺少 {column} 字段。"
            )

    levels: list[str] = []
    for raw_level in validation_issues["level"].tolist():
        level = _native_scalar(raw_level)
        if not isinstance(level, str) or level not in (
            "Fatal",
            "Error",
            "Warning",
        ):
            _raise_inconsistent_report_data("Validation Issues 包含无效 level。")
        levels.append(level)
    level_counts = Counter(levels)
    checks = (
        ("Validation", "Fatal Issues", level_counts["Fatal"]),
        ("Validation", "Error Issues", level_counts["Error"]),
        ("Validation", "Warning Issues", level_counts["Warning"]),
        (
            "Metrics",
            "Metrics Groups",
            0 if report_data.metrics is None else len(report_data.metrics),
        ),
        (
            "Diagnostics",
            "Diagnostic Issues",
            0 if report_data.diagnostics is None else len(report_data.diagnostics),
        ),
    )
    for section, item, actual in checks:
        expected = _summary_count(records, section=section, item=item)
        if expected != actual:
            _raise_inconsistent_report_data(
                f"Summary {section} / {item}={expected}，Detail 实际为 {actual}。"
            )

    code_checks = (
        (
            "Validation Code Counts",
            _detail_code_counts(
                validation_issues,
                detail_name="Validation Issues",
            ),
        ),
        (
            "Diagnostic Code Counts",
            _detail_code_counts(
                report_data.diagnostics,
                detail_name="Diagnostics",
            ),
        ),
    )
    for section, actual in code_checks:
        expected = _summary_code_counts(records, section=section)
        if expected != actual:
            _raise_inconsistent_report_data(
                f"Summary {section} 与 Detail Code Counts 不一致。"
            )


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, (str, bytes, bytearray, date, datetime, Enum)):
        return False
    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        return False
    try:
        return bool(missing)
    except (TypeError, ValueError):
        return False


def _native_scalar(value: Any) -> Any:
    if _is_missing(value):
        return None
    if isinstance(value, Enum):
        return value.value
    if hasattr(value, "item") and callable(value.item):
        try:
            return value.item()
        except (TypeError, ValueError, OverflowError):
            return value
    return value


def _excel_value(value: Any, *, force_date_text: bool = False) -> Any:
    value = _native_scalar(value)
    if value is None:
        return None
    if (
        isinstance(value, int)
        and not isinstance(value, bool)
        and abs(value) > _EXCEL_EXACT_INTEGER_MAX
    ):
        return str(value)
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        if force_date_text:
            return value.isoformat()
        if _EXCEL_NATIVE_DATE_MIN <= value <= _EXCEL_NATIVE_DATE_MAX:
            return value
        return value.isoformat()
    return value


def _date_columns_requiring_text(dataframe: pd.DataFrame) -> frozenset[str]:
    fallback_columns: set[str] = set()
    for column in dataframe.columns:
        if str(column) != "date":
            continue
        for raw_value in dataframe[column].tolist():
            value = _native_scalar(raw_value)
            if isinstance(value, datetime):
                value = value.date()
            if isinstance(value, date) and not (
                _EXCEL_NATIVE_DATE_MIN <= value <= _EXCEL_NATIVE_DATE_MAX
            ):
                fallback_columns.add(str(column))
                break
    return frozenset(fallback_columns)


def _json_safe(
    value: Any,
    *,
    depth: int = 0,
    ancestors: set[int] | None = None,
) -> Any:
    value = _native_scalar(value)
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (dict, list, tuple, set, frozenset)):
        if depth >= MAX_EVIDENCE_DEPTH:
            raise ReportError(
                INVALID_REPORT_DATA,
                f"Evidence 嵌套深度超过上限 {MAX_EVIDENCE_DEPTH}。",
            )
        if ancestors is None:
            ancestors = set()
        identity = id(value)
        if identity in ancestors:
            raise ReportError(
                INVALID_REPORT_DATA,
                "Evidence 包含循环引用。",
            )
        ancestors.add(identity)
        try:
            if isinstance(value, dict):
                return {
                    str(key): _json_safe(
                        item,
                        depth=depth + 1,
                        ancestors=ancestors,
                    )
                    for key, item in value.items()
                }
            normalized = [
                _json_safe(
                    item,
                    depth=depth + 1,
                    ancestors=ancestors,
                )
                for item in value
            ]
            if isinstance(value, (set, frozenset)):
                return sorted(normalized, key=repr)
            return normalized
        finally:
            ancestors.remove(identity)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _serialize_evidence(value: Any) -> str | None:
    if _is_missing(value):
        return None
    return json.dumps(
        _json_safe(value),
        ensure_ascii=False,
        sort_keys=True,
        allow_nan=False,
    )


def _set_cell_value(cell: Cell, value: Any) -> None:
    if isinstance(value, str) and len(value) > EXCEL_MAX_CELL_TEXT_LENGTH:
        raise ReportError(
            EXCEL_CELL_TEXT_TOO_LONG,
            (
                f"Sheet={cell.parent.title!r}, row={cell.row}, "
                f"column={cell.column_letter}: 文本长度 {len(value)} 超过 "
                f"Excel 上限 {EXCEL_MAX_CELL_TEXT_LENGTH}。"
            ),
        )
    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"


def _style_header(worksheet: Worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 22


def _write_dataframe(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    text_date_columns = _date_columns_requiring_text(dataframe)
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        _set_cell_value(worksheet.cell(row=1, column=column_index), str(column_name))
    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            column_name = str(dataframe.columns[column_index - 1])
            cell_value = (
                _serialize_evidence(value)
                if column_name == "evidence"
                else _excel_value(
                    value,
                    force_date_text=column_name in text_date_columns,
                )
            )
            _set_cell_value(
                worksheet.cell(row=row_index, column=column_index),
                cell_value,
            )
    _style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _write_not_generated(worksheet: Worksheet) -> None:
    _set_cell_value(worksheet.cell(row=1, column=1), VALIDATION_FAILED_MESSAGE)
    worksheet["A1"].font = Font(bold=True)
    worksheet["A1"].fill = _NOTE_FILL
    worksheet["A1"].alignment = Alignment(wrap_text=True, vertical="top")


def _apply_date_formats(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        if column_name != "date":
            continue
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = DATE_FORMAT


def _apply_metrics_formats(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        if column_name in COUNT_COLUMNS:
            number_format = COUNT_FORMAT
        elif column_name in USD_COLUMNS:
            number_format = USD_FORMAT
        elif column_name in PERCENTAGE_COLUMNS:
            number_format = PERCENTAGE_FORMAT
        elif column_name == "roas":
            number_format = ROAS_FORMAT
        else:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.number_format = number_format


def _diagnostic_number_format(metric: Any) -> str | None:
    if metric in PERCENTAGE_COLUMNS:
        return PERCENTAGE_FORMAT
    if metric == "roas":
        return ROAS_FORMAT
    if metric in COUNT_COLUMNS:
        return COUNT_FORMAT
    if metric in USD_COLUMNS:
        return USD_FORMAT
    return None


def _apply_diagnostics_formats(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
) -> None:
    columns = {
        str(name): position for position, name in enumerate(dataframe.columns, 1)
    }
    metric_column = columns.get("metric")
    if metric_column is None:
        return
    formatted_columns = [
        position
        for name in ("actual_value", "threshold")
        if (position := columns.get(name)) is not None
    ]
    for row_index in range(2, worksheet.max_row + 1):
        number_format = _diagnostic_number_format(
            worksheet.cell(row=row_index, column=metric_column).value
        )
        if number_format is None:
            continue
        for column_index in formatted_columns:
            worksheet.cell(row=row_index, column=column_index).number_format = (
                number_format
            )


def _apply_summary_formats(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    if "Value" not in dataframe.columns:
        return
    value_column = dataframe.columns.get_loc("Value") + 1
    for row_index in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_index, column=value_column).value
        if isinstance(value, int) and not isinstance(value, bool):
            worksheet.cell(row=row_index, column=value_column).number_format = (
                COUNT_FORMAT
            )


def _apply_alignment_and_widths(worksheet: Worksheet) -> None:
    header_by_column = {
        column_index: str(worksheet.cell(row=1, column=column_index).value or "")
        for column_index in range(1, worksheet.max_column + 1)
    }
    for column_index in range(1, worksheet.max_column + 1):
        header = header_by_column[column_index]
        max_length = len(header)
        for row_index in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            text_value = "" if cell.value is None else str(cell.value)
            line_lengths = [len(line) for line in text_value.splitlines()] or [0]
            max_length = max(max_length, *line_lengths)
            if header in _LONG_TEXT_COLUMNS and row_index > 1:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        minimum = 30 if header in _LONG_TEXT_COLUMNS else _MIN_COLUMN_WIDTH
        width = min(_MAX_COLUMN_WIDTH, max(minimum, max_length + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.sheet_view.showGridLines = False


def _validate_excel_row_limit(
    dataframe: pd.DataFrame,
    *,
    sheet_name: str,
) -> None:
    row_count = len(dataframe)
    if row_count > EXCEL_MAX_DATA_ROWS:
        raise ReportError(
            EXCEL_ROW_LIMIT_EXCEEDED,
            (
                f"Sheet={sheet_name!r}: 数据行数 {row_count} 超过 Excel 上限 "
                f"{EXCEL_MAX_DATA_ROWS}（另含 1 行 Header）。"
            ),
        )


def _validate_excel_row_limits(report_data: ReportData) -> None:
    tables = (
        (SHEET_NAMES[1], report_data.validation_issues),
        (SHEET_NAMES[2], report_data.metrics),
        (SHEET_NAMES[3], report_data.diagnostics),
    )
    for sheet_name, dataframe in tables:
        if dataframe is not None:
            _validate_excel_row_limit(dataframe, sheet_name=sheet_name)


def _build_workbook(report_data: ReportData) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = SHEET_NAMES[0]
    validation_sheet = workbook.create_sheet(SHEET_NAMES[1])
    metrics_sheet = workbook.create_sheet(SHEET_NAMES[2])
    diagnostics_sheet = workbook.create_sheet(SHEET_NAMES[3])

    _write_dataframe(summary_sheet, report_data.summary)
    _apply_date_formats(summary_sheet, report_data.summary)
    _apply_summary_formats(summary_sheet, report_data.summary)
    _write_dataframe(validation_sheet, report_data.validation_issues)
    _apply_date_formats(validation_sheet, report_data.validation_issues)

    if report_data.metrics is None:
        _write_not_generated(metrics_sheet)
    else:
        _write_dataframe(metrics_sheet, report_data.metrics)
        _apply_date_formats(metrics_sheet, report_data.metrics)
        _apply_metrics_formats(metrics_sheet, report_data.metrics)

    if report_data.diagnostics is None:
        _write_not_generated(diagnostics_sheet)
    else:
        _write_dataframe(diagnostics_sheet, report_data.diagnostics)
        _apply_date_formats(diagnostics_sheet, report_data.diagnostics)
        _apply_diagnostics_formats(diagnostics_sheet, report_data.diagnostics)

    for worksheet in workbook.worksheets:
        _apply_alignment_and_widths(worksheet)
    return workbook


def generate_excel_report(report_data: ReportData) -> bytes:
    """Serialize ReportData as a valid XLSX workbook held entirely in memory."""

    data = _validate_report_data(report_data)
    try:
        _validate_excel_row_limits(data)
        _validate_report_consistency(data)
        workbook = _build_workbook(data)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    except ReportError:
        raise
    except Exception as exc:
        raise ReportError(
            EXCEL_EXPORT_ERROR,
            "Excel 报告生成失败。",
        ) from exc
