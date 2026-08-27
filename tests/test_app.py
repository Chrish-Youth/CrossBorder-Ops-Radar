from __future__ import annotations

import csv
from io import BytesIO, StringIO
import importlib
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
import pytest
from streamlit.testing.v1 import AppTest

import app
import src.pipeline as pipeline_module
import src.report as report_module
from src.config import REQUIRED_COLUMNS
from src.diagnostics import DiagnosticsError
from src.loader import DataLoadError
from src.metrics import MetricsCalculationError
from src.pipeline import PipelineError, PipelineStatus, run_pipeline
from src.report import ReportError, VALIDATION_FAILED_MESSAGE


PROJECT_ROOT = Path(__file__).parents[1]
APP_PATH = PROJECT_ROOT / "app.py"
SAMPLE_PATH = PROJECT_ROOT / "data" / "sample_ecommerce_data.csv"


def make_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "date": "2026-08-24",
        "marketplace": "Amazon",
        "country": "US",
        "sku": "SKU-A",
        "product_name": "Example Product",
        "impressions": 2000,
        "clicks": 100,
        "orders": 10,
        "units_sold": 10,
        "sales": 200.0,
        "ad_spend": 50.0,
        "refunds": 0,
        "inventory": 20,
    }
    row.update(overrides)
    return row


def csv_content(
    *rows: dict[str, object],
    columns: tuple[str, ...] = REQUIRED_COLUMNS,
) -> bytes:
    text = StringIO(newline="")
    writer = csv.DictWriter(text, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return text.getvalue().encode("utf-8")


def xlsx_content(*rows: dict[str, object]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False)
    return output.getvalue()


def app_test() -> AppTest:
    return AppTest.from_file(str(APP_PATH), default_timeout=20).run()


def upload_and_run(
    filename: str,
    content: bytes,
    *,
    mime_type: str = "text/csv",
    group_label: str = "SKU",
) -> AppTest:
    at = app_test()
    at.file_uploader[0].upload(filename, content, mime_type).run()
    at.selectbox[0].select(group_label).run()
    return at.button[0].click().run(timeout=20)


def workbook_from_session(at: AppTest):
    excel_bytes = at.session_state["excel_bytes"]
    assert excel_bytes is not None
    return load_workbook(BytesIO(excel_bytes))


def workbook_summary_values(workbook: object) -> dict[tuple[str, str], object]:
    worksheet = workbook["Summary"]  # type: ignore[index]
    return {
        (row[0].value, row[1].value): row[2].value
        for row in worksheet.iter_rows(min_row=2)
    }


@pytest.mark.parametrize(
    ("label", "expected"),
    [
        ("Overall", None),
        ("SKU", ["sku"]),
        ("Marketplace", ["marketplace"]),
        ("Country", ["country"]),
        ("Marketplace + Country", ["marketplace", "country"]),
        (
            "Marketplace + Country + SKU",
            ["marketplace", "country", "sku"],
        ),
        (
            "Date + Marketplace + Country + SKU",
            ["date", "marketplace", "country", "sku"],
        ),
    ],
)
def test_group_by_options_map_to_metrics_api(
    label: str,
    expected: list[str] | None,
) -> None:
    assert app.resolve_group_by(label) == expected


def test_group_by_resolution_returns_an_independent_list() -> None:
    first = app.resolve_group_by("SKU")
    assert first is not None
    first.append("country")

    assert app.resolve_group_by("SKU") == ["sku"]


def test_unknown_group_by_label_is_rejected() -> None:
    with pytest.raises(ValueError):
        app.resolve_group_by("By Revenue")


@pytest.mark.parametrize(
    ("filename", "expected"),
    [
        ("amazon.csv", "amazon_crossborder_ops_radar.xlsx"),
        ("amazon.data.csv", "amazon.data_crossborder_ops_radar.xlsx"),
        ("archive.tar.csv", "archive.tar_crossborder_ops_radar.xlsx"),
        ("../../amazon.csv", "amazon_crossborder_ops_radar.xlsx"),
        (r"..\..\amazon.xlsx", "amazon_crossborder_ops_radar.xlsx"),
        ("商品 数据.csv", "商品_数据_crossborder_ops_radar.xlsx"),
        (".", app.DEFAULT_DOWNLOAD_FILENAME),
        ("..", app.DEFAULT_DOWNLOAD_FILENAME),
        ("", app.DEFAULT_DOWNLOAD_FILENAME),
        (None, app.DEFAULT_DOWNLOAD_FILENAME),
    ],
)
def test_download_filename_is_deterministic_and_path_safe(
    filename: str | None,
    expected: str,
) -> None:
    first = app.build_download_filename(filename)
    second = app.build_download_filename(filename)

    assert first == second == expected
    assert len(first.encode("utf-8")) <= app.MAX_DOWNLOAD_FILENAME_BYTES


@pytest.mark.parametrize("extra_bytes", [0, 1, 5000])
def test_ascii_download_filename_respects_utf8_byte_boundary(
    extra_bytes: int,
) -> None:
    suffix_bytes = len(app.DOWNLOAD_FILENAME_SUFFIX.encode("utf-8"))
    stem_budget = app.MAX_DOWNLOAD_FILENAME_BYTES - suffix_bytes
    stem = "a" * (stem_budget + extra_bytes)

    filename = app.build_download_filename(f"{stem}.csv")

    assert filename == "a" * stem_budget + app.DOWNLOAD_FILENAME_SUFFIX
    assert len(filename.encode("utf-8")) == app.MAX_DOWNLOAD_FILENAME_BYTES


def test_unicode_download_filename_truncates_on_character_boundary() -> None:
    source_stem = "商品数据" * 1000

    filename = app.build_download_filename(f"{source_stem}.csv")
    encoded = filename.encode("utf-8")
    truncated_stem = filename.removesuffix(app.DOWNLOAD_FILENAME_SUFFIX)

    assert len(encoded) <= app.MAX_DOWNLOAD_FILENAME_BYTES
    assert encoded.decode("utf-8") == filename
    assert filename.endswith(app.DOWNLOAD_FILENAME_SUFFIX)
    assert source_stem.startswith(truncated_stem)
    assert "�" not in filename
    assert app.build_download_filename(f"{source_stem}.csv") == filename


def test_default_download_filename_respects_byte_limit() -> None:
    assert (
        len(app.DEFAULT_DOWNLOAD_FILENAME.encode("utf-8"))
        <= app.MAX_DOWNLOAD_FILENAME_BYTES
    )


def test_analysis_signature_is_stable_and_sensitive_to_content_and_grain() -> None:
    content = b"same-file-content"
    first = app.build_analysis_signature(content, "input.csv", ["sku"])
    second = app.build_analysis_signature(content, "input.csv", ["sku"])
    changed_content = app.build_analysis_signature(
        b"different-file-content",
        "input.csv",
        ["sku"],
    )
    changed_filename = app.build_analysis_signature(
        content,
        "renamed.csv",
        ["sku"],
    )
    changed_grain = app.build_analysis_signature(content, "input.csv", None)

    assert first == second
    assert first != changed_content
    assert first != changed_filename
    assert first != changed_grain


def test_analysis_signature_freezes_group_by_representation() -> None:
    content = b"signature-content"

    overall_none = app.build_analysis_signature(content, "input.csv", None)
    overall_list = app.build_analysis_signature(content, "input.csv", [])
    overall_tuple = app.build_analysis_signature(  # type: ignore[arg-type]
        content,
        "input.csv",
        (),
    )
    sku_list = app.build_analysis_signature(content, "input.csv", ["sku"])
    sku_tuple = app.build_analysis_signature(  # type: ignore[arg-type]
        content,
        "input.csv",
        ("sku",),
    )

    assert overall_none != overall_list
    assert overall_list == overall_tuple
    assert sku_list == sku_tuple


@pytest.mark.parametrize(
    ("error", "title", "code", "stage"),
    [
        (
            DataLoadError("EMPTY_FILE", "empty"),
            "File could not be loaded.",
            "EMPTY_FILE",
            None,
        ),
        (
            PipelineError("INVALID_STAGE_RESULT", "metrics", "invalid"),
            "Internal pipeline contract error.",
            "INVALID_STAGE_RESULT",
            "metrics",
        ),
        (
            MetricsCalculationError("INVALID_GROUP_BY", "invalid"),
            "Metrics calculation failed.",
            "INVALID_GROUP_BY",
            None,
        ),
        (
            DiagnosticsError("INVALID_DIAGNOSTIC_INPUT", "invalid"),
            "Diagnostics failed.",
            "INVALID_DIAGNOSTIC_INPUT",
            None,
        ),
        (
            ReportError("EXCEL_ROW_LIMIT_EXCEEDED", "too many rows"),
            "Excel report could not be generated.",
            "EXCEL_ROW_LIMIT_EXCEEDED",
            None,
        ),
    ],
)
def test_structured_errors_have_safe_ui_presentations(
    error: BaseException,
    title: str,
    code: str,
    stage: str | None,
) -> None:
    presentation = app.error_presentation(error)

    assert presentation.title == title
    assert presentation.code == code
    assert presentation.stage == stage


def test_unexpected_error_does_not_leak_exception_text() -> None:
    presentation = app.error_presentation(RuntimeError("secret traceback detail"))

    assert presentation.title == "Unexpected application error."
    assert presentation.code == "UNEXPECTED_APPLICATION_ERROR"
    assert "secret" not in presentation.message


def test_metrics_display_formats_values_without_mutating_source() -> None:
    metrics = pd.DataFrame(
        {
            "sku": ["NAN", "ZERO", "VALUE"],
            "ctr": [float("nan"), 0.0, 0.0125],
            "gmv": [float("nan"), 0.0, 1234.5],
            "roas": [float("nan"), 0.0, 2.5],
        }
    )
    before = metrics.copy(deep=True)

    display = app.build_metrics_display(metrics)

    assert display["ctr"].tolist() == ["—", "0.00%", "1.25%"]
    assert display["gmv"].tolist() == ["—", "$0.00", "$1,234.50"]
    assert display["roas"].tolist() == ["—", "0.00x", "2.50x"]
    pd.testing.assert_frame_equal(metrics, before)


def test_diagnostics_display_preserves_all_rows_and_source_evidence() -> None:
    result = run_pipeline(SAMPLE_PATH, group_by="sku")
    assert result.diagnostics is not None
    before = result.diagnostics.copy(deep=True)
    evidence_before = result.diagnostics.loc[0, "evidence"].copy()

    display = app.build_diagnostics_display(result.diagnostics)

    assert len(display) == len(result.diagnostics) == 11
    assert display.loc[0, "evidence"] == (
        '{"minimum_orders": 10, "orders": 20}'
    )
    no_order_codes = display.loc[
        display["sku"] == "SKU-NO-ORDER",
        "code",
    ].tolist()
    assert no_order_codes == [
        "LOW_CVR",
        "CLICKS_WITHOUT_ORDERS",
        "SPEND_WITHOUT_ORDERS",
        "LOW_ROAS",
    ]
    pd.testing.assert_frame_equal(result.diagnostics, before)
    assert result.diagnostics.loc[0, "evidence"] == evidence_before


def test_execute_analysis_runs_real_sample_and_builds_excel_once() -> None:
    artifacts = app.execute_analysis(
        SAMPLE_PATH.read_bytes(),
        filename=SAMPLE_PATH.name,
        group_by=["sku"],
    )

    assert artifacts.pipeline_result.status is PipelineStatus.SUCCESS
    assert artifacts.pipeline_result.validation.report.total_rows == 23
    assert artifacts.pipeline_result.validation.report.valid_rows == 14
    assert artifacts.pipeline_result.metrics is not None
    assert len(artifacts.pipeline_result.metrics) == 12
    assert artifacts.pipeline_result.diagnostics is not None
    assert len(artifacts.pipeline_result.diagnostics) == 11
    assert artifacts.report_data is not None
    assert artifacts.report_error is None
    assert artifacts.excel_bytes is not None
    assert artifacts.excel_bytes.startswith(b"PK")


def test_execute_analysis_preserves_pipeline_when_report_fails(
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
) -> None:
    failure = ReportError("EXCEL_CELL_TEXT_TOO_LONG", "too long")

    def fail_report(_: object) -> bytes:
        raise failure

    monkeypatch.setattr(app, "generate_excel_report", fail_report)

    artifacts = app.execute_analysis(
        SAMPLE_PATH.read_bytes(),
        filename=SAMPLE_PATH.name,
        group_by=["sku"],
    )

    assert artifacts.pipeline_result.metrics is not None
    assert len(artifacts.pipeline_result.metrics) == 12
    assert artifacts.pipeline_result.diagnostics is not None
    assert len(artifacts.pipeline_result.diagnostics) == 11
    assert artifacts.report_data is not None
    assert artifacts.excel_bytes is None
    assert artifacts.report_error is not None
    assert artifacts.report_error.code == "EXCEL_CELL_TEXT_TOO_LONG"
    assert not any(
        record.getMessage() == "Unexpected error during Excel report generation"
        for record in caplog.records
    )


def test_unexpected_report_exception_is_logged_without_breaking_pipeline(
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
) -> None:
    def fail_report(_: object) -> bytes:
        raise RuntimeError("private report detail")

    monkeypatch.setattr(app, "generate_excel_report", fail_report)
    caplog.set_level(logging.ERROR)

    artifacts = app.execute_analysis(
        SAMPLE_PATH.read_bytes(),
        filename=SAMPLE_PATH.name,
        group_by=["sku"],
    )

    assert artifacts.pipeline_result.metrics is not None
    assert artifacts.pipeline_result.diagnostics is not None
    assert artifacts.excel_bytes is None
    assert artifacts.report_error is not None
    assert artifacts.report_error.code == "UNEXPECTED_REPORT_ERROR"
    records = [
        record
        for record in caplog.records
        if record.getMessage()
        == "Unexpected error during Excel report generation"
    ]
    assert len(records) == 1
    assert records[0].exc_info is not None
    assert isinstance(records[0].exc_info[1], RuntimeError)


def test_import_app_does_not_run_pipeline(monkeypatch: pytest.MonkeyPatch) -> None:
    calls: list[object] = []
    original_run_pipeline = app.run_pipeline

    def must_not_run(*args: object, **kwargs: object) -> None:
        calls.append((args, kwargs))

    monkeypatch.setattr(pipeline_module, "run_pipeline", must_not_run)

    try:
        importlib.reload(app)
        assert calls == []
    finally:
        app.run_pipeline = original_run_pipeline


def test_initial_ui_waits_for_upload_and_does_not_run_automatically() -> None:
    at = app_test()

    assert list(at.exception) == []
    assert at.title[0].value == app.APP_TITLE
    assert at.selectbox[0].value == app.DEFAULT_GROUP_BY_LABEL
    assert at.button[0].disabled is True
    assert [message.value for message in at.info] == [
        "Upload a CSV or XLSX file to begin."
    ]
    assert len(at.dataframe) == 0
    assert len(at.download_button) == 0


def test_uploaded_file_is_ready_but_not_analyzed_until_button_click() -> None:
    at = app_test()
    at.file_uploader[0].upload(
        SAMPLE_PATH.name,
        SAMPLE_PATH.read_bytes(),
        "text/csv",
    ).run()

    assert at.button[0].disabled is False
    assert [message.value for message in at.info] == ["Ready to analyze."]
    assert len(at.dataframe) == 0
    assert at.session_state["pipeline_result"] is None


def test_sample_success_ui_shows_validation_metrics_diagnostics_and_download() -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())

    assert list(at.exception) == []
    assert [heading.value for heading in at.subheader] == [
        "Upload and configuration",
        "Validation",
        "Metrics",
        "Diagnostic Signals",
        "Excel Report",
    ]
    assert [(metric.label, metric.value) for metric in at.metric] == [
        ("Raw Rows", "23"),
        ("Valid Rows", "14"),
        ("Excluded Rows", "9"),
        ("Warning Rows", "3"),
        ("Fatal Issues", "0"),
        ("Error Issues", "8"),
        ("Warning Issues", "3"),
    ]
    assert [len(table.value) for table in at.dataframe] == [11, 12, 11]
    assert len(at.download_button) == 1
    assert at.session_state["excel_bytes"].startswith(b"PK")
    assert at.session_state["download_filename"] == (
        "sample_ecommerce_data_crossborder_ops_radar.xlsx"
    )
    assert at.session_state["report_error"] is None


def test_xlsx_upload_runs_through_pipeline_without_ui_format_sniffing() -> None:
    at = upload_and_run(
        "operations.xlsx",
        xlsx_content(make_row()),
        mime_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    assert list(at.exception) == []
    assert "Metrics" in [heading.value for heading in at.subheader]
    assert len(at.download_button) == 1
    assert at.session_state["pipeline_result"].metrics["sku"].tolist() == [
        "SKU-A"
    ]


def test_validation_failed_ui_hides_metrics_but_keeps_excel_download() -> None:
    columns = tuple(column for column in REQUIRED_COLUMNS if column != "sku")
    at = upload_and_run(
        "missing-sku.csv",
        csv_content(make_row(), columns=columns),
    )

    assert list(at.exception) == []
    assert at.session_state["pipeline_result"].status is PipelineStatus.VALIDATION_FAILED
    assert "File validation failed" in at.error[0].value
    assert [heading.value for heading in at.subheader] == [
        "Upload and configuration",
        "Validation",
        "Excel Report",
    ]
    assert len(at.dataframe) == 1
    assert at.dataframe[0].value["code"].tolist() == [
        "MISSING_REQUIRED_COLUMN"
    ]
    assert len(at.download_button) == 1


def test_loader_error_ui_is_structured_and_clears_results() -> None:
    at = upload_and_run(
        "broken.xlsx",
        b"not an xlsx workbook",
        mime_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    assert list(at.exception) == []
    assert [error.value for error in at.error] == ["File could not be loaded."]
    assert any("FILE_READ_ERROR" in markdown.value for markdown in at.markdown)
    assert at.session_state["pipeline_result"] is None
    assert len(at.dataframe) == 0
    assert len(at.download_button) == 0


def test_pipeline_error_ui_shows_code_and_stage_without_success(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_pipeline(*args: object, **kwargs: object) -> None:
        raise PipelineError("INVALID_STAGE_RESULT", "metrics", "invalid stage")

    monkeypatch.setattr(pipeline_module, "run_pipeline", fail_pipeline)
    at = upload_and_run("input.csv", csv_content(make_row()))

    assert list(at.exception) == []
    assert [error.value for error in at.error] == [
        "Internal pipeline contract error."
    ]
    assert any(
        "INVALID_STAGE_RESULT" in markdown.value and "metrics" in markdown.value
        for markdown in at.markdown
    )
    assert at.session_state["pipeline_result"] is None
    assert len(at.download_button) == 0


@pytest.mark.parametrize(
    ("failure", "expected_title", "expected_code"),
    [
        (
            MetricsCalculationError("COUNT_AGGREGATION_OVERFLOW", "overflow"),
            "Metrics calculation failed.",
            "COUNT_AGGREGATION_OVERFLOW",
        ),
        (
            DiagnosticsError("INVALID_DIAGNOSTIC_INPUT_VALUE", "invalid"),
            "Diagnostics failed.",
            "INVALID_DIAGNOSTIC_INPUT_VALUE",
        ),
        (
            RuntimeError("private traceback detail"),
            "Unexpected application error.",
            "UNEXPECTED_APPLICATION_ERROR",
        ),
    ],
)
def test_analysis_exception_ui_is_structured_without_partial_success(
    failure: Exception,
    expected_title: str,
    expected_code: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_pipeline(*args: object, **kwargs: object) -> None:
        raise failure

    monkeypatch.setattr(pipeline_module, "run_pipeline", fail_pipeline)
    at = upload_and_run("input.csv", csv_content(make_row()))

    assert list(at.exception) == []
    assert [error.value for error in at.error] == [expected_title]
    assert any(expected_code in markdown.value for markdown in at.markdown)
    assert all(
        "private traceback detail" not in value
        for value in [
            *(error.value for error in at.error),
            *(markdown.value for markdown in at.markdown),
            *(caption.value for caption in at.caption),
        ]
    )
    assert at.session_state["pipeline_result"] is None
    assert len(at.download_button) == 0


def test_report_error_ui_keeps_analysis_and_disables_download(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_report(_: object) -> bytes:
        raise ReportError("EXCEL_ROW_LIMIT_EXCEEDED", "too many rows")

    monkeypatch.setattr(report_module, "generate_excel_report", fail_report)
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())

    assert list(at.exception) == []
    assert at.session_state["pipeline_result"].metrics is not None
    assert at.session_state["pipeline_result"].diagnostics is not None
    assert "Metrics" in [heading.value for heading in at.subheader]
    assert "Diagnostic Signals" in [heading.value for heading in at.subheader]
    assert any(
        error.value == "Excel report could not be generated."
        for error in at.error
    )
    assert any(
        "EXCEL_ROW_LIMIT_EXCEEDED" in markdown.value
        for markdown in at.markdown
    )
    assert len(at.download_button) == 0


def test_success_then_loader_failure_does_not_leak_previous_results() -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())
    assert at.session_state["pipeline_result"] is not None
    assert at.session_state["excel_bytes"] is not None

    at.file_uploader[0].upload(
        "broken.xlsx",
        b"not an xlsx workbook",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).run()

    assert at.session_state["pipeline_result"] is None
    assert at.session_state["report_data"] is None
    assert at.session_state["excel_bytes"] is None
    assert at.session_state["report_error"] is None
    assert len(at.dataframe) == 0
    assert len(at.download_button) == 0

    at.button[0].click().run(timeout=20)

    assert [error.value for error in at.error] == ["File could not be loaded."]
    assert at.session_state["pipeline_result"] is None
    assert at.session_state["report_data"] is None
    assert at.session_state["excel_bytes"] is None
    assert at.session_state["report_error"] is None
    assert len(at.dataframe) == 0
    assert len(at.download_button) == 0
    assert "Metrics" not in [heading.value for heading in at.subheader]
    assert "Diagnostic Signals" not in [
        heading.value for heading in at.subheader
    ]


def test_success_then_pipeline_error_and_same_session_recovery(
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
) -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())
    original_run_pipeline = pipeline_module.run_pipeline

    def fail_pipeline(*args: object, **kwargs: object) -> None:
        raise PipelineError("RECOVERY_PIPELINE_ERROR", "metrics", "known failure")

    monkeypatch.setattr(pipeline_module, "run_pipeline", fail_pipeline)
    caplog.set_level(logging.ERROR)
    at.button[0].click().run(timeout=20)

    assert at.session_state["pipeline_result"] is None
    assert at.session_state["report_data"] is None
    assert at.session_state["excel_bytes"] is None
    assert at.session_state["report_error"] is None
    assert at.session_state["analysis_error"].code == "RECOVERY_PIPELINE_ERROR"
    assert len(at.dataframe) == 0
    assert len(at.download_button) == 0
    assert not any(
        record.getMessage() == "Unexpected application error during analysis"
        for record in caplog.records
    )

    monkeypatch.setattr(
        pipeline_module,
        "run_pipeline",
        original_run_pipeline,
    )
    at.button[0].click().run(timeout=20)

    assert list(at.exception) == []
    assert at.session_state["analysis_error"] is None
    assert at.session_state["report_error"] is None
    assert at.session_state["pipeline_result"].status is PipelineStatus.SUCCESS
    assert len(at.session_state["pipeline_result"].metrics) == 12
    assert len(at.session_state["pipeline_result"].diagnostics) == 11
    assert at.session_state["excel_bytes"].startswith(b"PK")
    assert len(at.download_button) == 1
    assert not any("RECOVERY_PIPELINE_ERROR" in error.value for error in at.error)


def test_unexpected_error_logs_traceback_and_clears_previous_success(
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
) -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())

    def fail_pipeline(*args: object, **kwargs: object) -> None:
        raise RuntimeError("private developer detail")

    monkeypatch.setattr(pipeline_module, "run_pipeline", fail_pipeline)
    caplog.set_level(logging.ERROR)
    at.button[0].click().run(timeout=20)

    assert at.session_state["pipeline_result"] is None
    assert at.session_state["report_data"] is None
    assert at.session_state["excel_bytes"] is None
    assert at.session_state["report_error"] is None
    assert at.session_state["analysis_error"].code == (
        "UNEXPECTED_APPLICATION_ERROR"
    )
    assert len(at.dataframe) == 0
    assert len(at.download_button) == 0
    ui_text = [
        *(error.value for error in at.error),
        *(markdown.value for markdown in at.markdown),
        *(caption.value for caption in at.caption),
    ]
    assert all("private developer detail" not in value for value in ui_text)
    assert all("RuntimeError" not in value for value in ui_text)
    assert all("Traceback" not in value for value in ui_text)
    records = [
        record
        for record in caplog.records
        if record.getMessage() == "Unexpected application error during analysis"
    ]
    assert len(records) == 1
    assert records[0].exc_info is not None
    assert isinstance(records[0].exc_info[1], RuntimeError)


def test_report_error_then_same_input_success_clears_error_and_restores_download(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    original_generate_excel_report = report_module.generate_excel_report

    def fail_report(_: object) -> bytes:
        raise ReportError("RECOVERY_REPORT_ERROR", "known report failure")

    monkeypatch.setattr(report_module, "generate_excel_report", fail_report)
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())

    assert at.session_state["pipeline_result"].status is PipelineStatus.SUCCESS
    assert len(at.session_state["pipeline_result"].metrics) == 12
    assert len(at.session_state["pipeline_result"].diagnostics) == 11
    assert at.session_state["excel_bytes"] is None
    assert at.session_state["report_error"].code == "RECOVERY_REPORT_ERROR"
    assert len(at.download_button) == 0

    monkeypatch.setattr(
        report_module,
        "generate_excel_report",
        original_generate_excel_report,
    )
    at.button[0].click().run(timeout=20)

    assert list(at.exception) == []
    assert at.session_state["analysis_error"] is None
    assert at.session_state["report_error"] is None
    assert at.session_state["excel_bytes"].startswith(b"PK")
    assert len(at.download_button) == 1
    assert not any("RECOVERY_REPORT_ERROR" in error.value for error in at.error)


def test_same_uploaded_file_can_run_twice_without_pointer_consumption() -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())
    first_signature = at.session_state["analysis_signature"]
    first_metrics = at.session_state["pipeline_result"].metrics.copy(deep=True)
    first_diagnostics = at.session_state["pipeline_result"].diagnostics.copy(
        deep=True
    )
    assert at.session_state["excel_bytes"].startswith(b"PK")

    at.button[0].click().run(timeout=20)

    assert list(at.exception) == []
    assert at.session_state["analysis_signature"] == first_signature
    assert at.session_state["pipeline_result"].status is PipelineStatus.SUCCESS
    pd.testing.assert_frame_equal(
        at.session_state["pipeline_result"].metrics,
        first_metrics,
    )
    pd.testing.assert_frame_equal(
        at.session_state["pipeline_result"].diagnostics,
        first_diagnostics,
    )
    assert at.session_state["excel_bytes"].startswith(b"PK")
    assert len(at.download_button) == 1


def test_upload_change_invalidates_old_analysis_state() -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())
    assert len(at.dataframe) == 3
    old_signature = at.session_state["analysis_signature"]

    at.file_uploader[0].upload(
        "new.csv",
        csv_content(make_row(sku="SKU-NEW")),
        "text/csv",
    ).run()

    assert [message.value for message in at.info] == ["Ready to analyze."]
    assert len(at.dataframe) == 0
    assert len(at.download_button) == 0
    assert at.session_state["pipeline_result"] is None
    assert at.session_state["analysis_signature"] is None
    assert at.session_state["report_data"] is None
    assert at.session_state["excel_bytes"] is None
    assert at.session_state["report_error"] is None
    assert at.session_state["analysis_error"] is None
    assert at.session_state["download_filename"] is None
    assert old_signature is not None


def test_group_by_change_invalidates_old_analysis_state() -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())
    assert len(at.dataframe[1].value) == 12

    at.selectbox[0].select("Overall").run()

    assert [message.value for message in at.info] == ["Ready to analyze."]
    assert len(at.dataframe) == 0
    assert len(at.download_button) == 0
    assert at.session_state["pipeline_result"] is None
    assert at.session_state["analysis_signature"] is None
    assert at.session_state["report_data"] is None
    assert at.session_state["excel_bytes"] is None
    assert at.session_state["report_error"] is None
    assert at.session_state["analysis_error"] is None
    assert at.session_state["download_filename"] is None


def test_group_by_change_can_be_rerun_at_new_grain() -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())
    at.selectbox[0].select("Overall").run()
    at.button[0].click().run(timeout=20)

    assert list(at.exception) == []
    assert len(at.dataframe[1].value) == 1
    assert at.session_state["pipeline_result"].metrics.columns[0] == "impressions"


def test_group_by_change_replaces_sku_workbook_with_overall_workbook() -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())
    sku_bytes = at.session_state["excel_bytes"]
    sku_workbook = workbook_from_session(at)

    assert sku_workbook["Metrics"].max_row == 13
    assert sku_workbook["Metrics"]["A1"].value == "sku"

    at.selectbox[0].select("Overall").run()

    assert at.session_state["excel_bytes"] is None
    assert len(at.download_button) == 0
    assert len(at.dataframe) == 0
    assert [message.value for message in at.info] == ["Ready to analyze."]

    at.button[0].click().run(timeout=20)
    overall_workbook = workbook_from_session(at)
    overall_summary = workbook_summary_values(overall_workbook)

    assert at.session_state["excel_bytes"] != sku_bytes
    assert overall_workbook["Metrics"].max_row == 2
    assert overall_workbook["Metrics"]["A1"].value == "impressions"
    assert overall_summary[("Pipeline", "Status")] == "SUCCESS"
    assert overall_summary[("Metrics", "Metrics Groups")] == 1
    assert len(at.download_button) == 1


def test_success_workbook_is_replaced_by_current_fatal_workbook() -> None:
    at = upload_and_run(SAMPLE_PATH.name, SAMPLE_PATH.read_bytes())
    success_bytes = at.session_state["excel_bytes"]
    columns = tuple(column for column in REQUIRED_COLUMNS if column != "sku")

    at.file_uploader[0].upload(
        "fatal.csv",
        csv_content(make_row(), columns=columns),
        "text/csv",
    ).run()

    assert at.session_state["excel_bytes"] is None
    assert len(at.download_button) == 0

    at.button[0].click().run(timeout=20)
    fatal_workbook = workbook_from_session(at)
    fatal_summary = workbook_summary_values(fatal_workbook)
    validation_sheet = fatal_workbook["Validation Issues"]
    headers = [cell.value for cell in validation_sheet[1]]
    code_column = headers.index("code") + 1
    fatal_codes = [
        validation_sheet.cell(row=row, column=code_column).value
        for row in range(2, validation_sheet.max_row + 1)
    ]

    assert at.session_state["pipeline_result"].status is (
        PipelineStatus.VALIDATION_FAILED
    )
    assert at.session_state["excel_bytes"] != success_bytes
    assert fatal_summary[("Pipeline", "Status")] == "VALIDATION_FAILED"
    assert "MISSING_REQUIRED_COLUMN" in fatal_codes
    assert fatal_workbook["Metrics"]["A1"].value == VALIDATION_FAILED_MESSAGE
    assert (
        fatal_workbook["Diagnostics"]["A1"].value
        == VALIDATION_FAILED_MESSAGE
    )
    assert len(at.download_button) == 1
    assert "Metrics" not in [heading.value for heading in at.subheader]
    assert "Diagnostic Signals" not in [
        heading.value for heading in at.subheader
    ]


def test_lightweight_large_tables_render_all_rows_without_html_conversion() -> None:
    rows = [
        make_row(
            sku=f"SKU-{index:04d}",
            impressions=1000,
            clicks=0,
            orders=0,
            units_sold=0,
            sales=0,
            ad_spend=0,
            refunds=0,
            inventory=10,
        )
        for index in range(1000)
    ]

    at = upload_and_run("large.csv", csv_content(*rows))

    assert list(at.exception) == []
    assert [len(table.value) for table in at.dataframe] == [1000, 1000]
    assert len(at.session_state["pipeline_result"].metrics) == 1000
    assert len(at.session_state["pipeline_result"].diagnostics) == 1000
    assert at.session_state["excel_bytes"].startswith(b"PK")
    assert len(at.download_button) == 1


def test_no_diagnostics_ui_uses_explicit_empty_state() -> None:
    normal = make_row(
        impressions=100,
        clicks=10,
        orders=1,
        units_sold=1,
        sales=20.0,
        ad_spend=5.0,
        refunds=0,
        inventory=10,
    )
    at = upload_and_run("normal.csv", csv_content(normal))

    assert list(at.exception) == []
    assert any(
        "No diagnostic signals were triggered" in message.value
        for message in at.info
    )
    assert len(at.dataframe) == 1
    assert len(at.download_button) == 1
