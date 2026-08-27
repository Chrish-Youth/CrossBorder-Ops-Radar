from __future__ import annotations

from collections import Counter
import copy
import csv
from datetime import date, datetime
from io import BytesIO, StringIO
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import pytest

import src.report as report_module
from src.config import (
    CLICKS_WITHOUT_ORDERS,
    HIGH_IMPRESSIONS_LOW_CTR,
    HIGH_REFUND_RATE,
    LOW_CVR,
    LOW_ROAS,
    OUT_OF_STOCK,
    REQUIRED_COLUMNS,
    SPEND_WITHOUT_ORDERS,
)
from src.diagnostics import DIAGNOSTIC_ISSUE_COLUMNS
from src.metrics import BASE_MEASURES, DERIVED_METRICS
from src.pipeline import PipelineResult, PipelineStatus, run_pipeline
from src.report import (
    COUNT_FORMAT,
    DATE_FORMAT,
    EXCEL_CELL_TEXT_TOO_LONG,
    EXCEL_EXPORT_ERROR,
    EXCEL_MAX_CELL_TEXT_LENGTH,
    EXCEL_MAX_DATA_ROWS,
    EXCEL_ROW_LIMIT_EXCEEDED,
    INCONSISTENT_REPORT_DATA,
    INVALID_REPORT_DATA,
    INVALID_REPORT_INPUT,
    MAX_EVIDENCE_DEPTH,
    PERCENTAGE_FORMAT,
    ROAS_FORMAT,
    SHEET_NAMES,
    SUMMARY_COLUMNS,
    USD_FORMAT,
    VALIDATION_FAILED_MESSAGE,
    VALIDATION_ISSUE_COLUMNS,
    ReportData,
    ReportError,
    _validate_excel_row_limit,
    build_report_data,
    generate_excel_report,
)


SAMPLE_PATH = Path(__file__).parents[1] / "data" / "sample_ecommerce_data.csv"


def make_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "date": "2026-08-24",
        "marketplace": "Amazon",
        "country": "US",
        "sku": "SKU-A",
        "product_name": "Example Product",
        "impressions": 2000,
        "clicks": 100,
        "orders": 10,
        "units_sold": 10,
        "sales": 200.0,
        "ad_spend": 50.0,
        "refunds": 0,
        "inventory": 20,
    }
    row.update(overrides)
    return row


def csv_content(
    *rows: dict[str, object],
    columns: tuple[str, ...] = REQUIRED_COLUMNS,
) -> bytes:
    text = StringIO(newline="")
    writer = csv.DictWriter(text, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return text.getvalue().encode("utf-8")


def sample_result() -> PipelineResult:
    return run_pipeline(SAMPLE_PATH, group_by="sku")


def workbook_for(result: PipelineResult):
    report_data = build_report_data(result)
    excel_bytes = generate_excel_report(report_data)
    return report_data, excel_bytes, load_workbook(BytesIO(excel_bytes))


def summary_values(report_data: ReportData) -> dict[tuple[str, str], object]:
    return {
        (row.Section, row.Item): row.Value
        for row in report_data.summary.itertuples(index=False)
    }


def column_positions(worksheet) -> dict[str, int]:
    return {
        str(cell.value): position
        for position, cell in enumerate(worksheet[1], start=1)
    }


def nested_evidence(depth: int) -> object:
    value: object = "leaf"
    for _ in range(depth):
        value = {"child": value}
    return value


def evidence_with_serialized_length(target_length: int) -> dict[str, str]:
    empty_length = len(
        json.dumps(
            {"payload": ""},
            ensure_ascii=False,
            sort_keys=True,
            allow_nan=False,
        )
    )
    return {"payload": "x" * (target_length - empty_length)}


@pytest.mark.parametrize("invalid", [None, {}, pd.DataFrame()])
def test_build_report_data_rejects_non_pipeline_result(invalid: object) -> None:
    with pytest.raises(ReportError) as caught:
        build_report_data(invalid)  # type: ignore[arg-type]

    assert caught.value.code == INVALID_REPORT_INPUT


def test_build_report_data_rejects_inconsistent_pipeline_result() -> None:
    valid_result = sample_result()
    inconsistent = PipelineResult(
        status=PipelineStatus.SUCCESS,
        validation=valid_result.validation,
        metrics=None,
        diagnostics=None,
    )

    with pytest.raises(ReportError) as caught:
        build_report_data(inconsistent)

    assert caught.value.code == INVALID_REPORT_INPUT


def test_sample_report_summary_locks_real_pipeline_counts_and_codes() -> None:
    result = sample_result()
    report_data = build_report_data(result)
    values = summary_values(report_data)

    assert report_data.summary.columns.tolist() == list(SUMMARY_COLUMNS)
    assert values[("Pipeline", "Status")] == "SUCCESS"
    assert values[("Validation", "Raw Rows")] == 23
    assert values[("Validation", "Valid Rows")] == 14
    assert values[("Validation", "Excluded Rows")] == 9
    assert values[("Validation", "Warning Rows")] == 3
    assert values[("Validation", "Fatal Issues")] == 0
    assert values[("Validation", "Error Issues")] == 8
    assert values[("Validation", "Warning Issues")] == 3
    assert values[("Metrics", "Metrics Groups")] == 12
    assert values[("Diagnostics", "Diagnostic Issues")] == 11

    validation_counts = Counter(issue.code for issue in result.validation.report.issues)
    diagnostic_counts = Counter(result.diagnostics["code"].tolist())  # type: ignore[index]
    assert {
        item: value
        for (section, item), value in values.items()
        if section == "Validation Code Counts"
    } == validation_counts
    assert {
        item: value
        for (section, item), value in values.items()
        if section == "Diagnostic Code Counts"
    } == diagnostic_counts


def test_validation_issues_preserve_report_order_and_schema() -> None:
    result = sample_result()
    report_data = build_report_data(result)

    assert report_data.validation_issues.columns.tolist() == list(
        VALIDATION_ISSUE_COLUMNS
    )
    assert report_data.validation_issues["code"].tolist() == [
        issue.code for issue in result.validation.report.issues
    ]
    assert report_data.validation_issues["level"].tolist() == [
        issue.level for issue in result.validation.report.issues
    ]
    assert str(report_data.validation_issues["row"].dtype) == "Int64"


def test_report_data_is_independent_and_export_does_not_mutate_pipeline() -> None:
    result = sample_result()
    clean_before = result.validation.clean_data.copy(deep=True)
    metrics_before = result.metrics.copy(deep=True)  # type: ignore[union-attr]
    diagnostics_before = result.diagnostics.copy(deep=True)  # type: ignore[union-attr]
    evidence_before = copy.deepcopy(result.diagnostics.loc[0, "evidence"])  # type: ignore[union-attr]

    report_data = build_report_data(result)
    report_data.metrics.loc[0, "gmv"] = -1  # type: ignore[union-attr]
    evidence = report_data.diagnostics.loc[0, "evidence"]  # type: ignore[union-attr]
    evidence[next(iter(evidence))] = -1
    generate_excel_report(report_data)

    pd.testing.assert_frame_equal(result.validation.clean_data, clean_before)
    pd.testing.assert_frame_equal(result.metrics, metrics_before)
    pd.testing.assert_frame_equal(result.diagnostics, diagnostics_before)
    assert result.diagnostics.loc[0, "evidence"] == evidence_before  # type: ignore[union-attr]


def test_workbook_has_fixed_structure_and_table_usability() -> None:
    _, excel_bytes, workbook = workbook_for(sample_result())

    assert len(excel_bytes) > 0
    assert workbook.sheetnames == list(SHEET_NAMES)
    for sheet_name in SHEET_NAMES:
        worksheet = workbook[sheet_name]
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == worksheet.dimensions
        assert all(cell.font.bold for cell in worksheet[1])
        assert worksheet.sheet_view.showGridLines is False


def test_sample_metrics_values_and_formats_match_metrics_dataframe() -> None:
    result = sample_result()
    report_data, _, workbook = workbook_for(result)
    worksheet = workbook["Metrics"]
    columns = column_positions(worksheet)
    sku_row = next(
        row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=columns["sku"]).value == "SKU-NORMAL-US"
    )

    metrics_row = report_data.metrics.loc[  # type: ignore[union-attr]
        report_data.metrics["sku"] == "SKU-NORMAL-US"  # type: ignore[union-attr]
    ].iloc[0]
    expected_values = {
        "gmv": 1979.34,
        "ctr": 0.03,
        "roas": 1979.34 / 310.0,
        "inventory": 95,
    }
    for column, expected in expected_values.items():
        cell = worksheet.cell(row=sku_row, column=columns[column])
        assert cell.value == pytest.approx(expected)
        assert cell.value == pytest.approx(metrics_row[column])

    assert worksheet.cell(sku_row, columns["gmv"]).number_format == USD_FORMAT
    assert worksheet.cell(sku_row, columns["ctr"]).number_format == PERCENTAGE_FORMAT
    assert worksheet.cell(sku_row, columns["roas"]).number_format == ROAS_FORMAT
    assert worksheet.cell(sku_row, columns["inventory"]).number_format == COUNT_FORMAT


def test_nan_is_blank_but_valid_zero_remains_numeric() -> None:
    result = run_pipeline(
        csv_content(
            make_row(
                sku="SKU-NAN",
                impressions=0,
                clicks=0,
                orders=0,
                units_sold=0,
                sales=0,
                ad_spend=0,
                inventory=0,
            ),
            make_row(
                sku="SKU-ZERO",
                impressions=100,
                clicks=0,
                orders=0,
                units_sold=0,
                sales=0,
                ad_spend=10,
                inventory=0,
            ),
        ),
        filename="zero.csv",
        group_by="sku",
    )
    report_data, _, workbook = workbook_for(result)
    worksheet = workbook["Metrics"]
    columns = column_positions(worksheet)
    rows = {
        worksheet.cell(row=row, column=columns["sku"]).value: row
        for row in range(2, worksheet.max_row + 1)
    }

    assert pd.isna(
        report_data.metrics.loc[report_data.metrics["sku"] == "SKU-NAN", "ctr"].iloc[0]  # type: ignore[union-attr]
    )
    nan_cell = worksheet.cell(rows["SKU-NAN"], columns["ctr"])
    zero_cell = worksheet.cell(rows["SKU-ZERO"], columns["ctr"])
    zero_money_cell = worksheet.cell(rows["SKU-ZERO"], columns["gmv"])
    assert nan_cell.value is None
    assert zero_cell.value == 0
    assert zero_cell.number_format == PERCENTAGE_FORMAT
    assert zero_money_cell.value == 0
    assert zero_money_cell.number_format == USD_FORMAT


def test_large_integer_fallback_is_value_based_and_lossless() -> None:
    boundary_values = (
        2**53 - 1,
        2**53,
        2**53 + 1,
        9007199254740993,
    )
    result = run_pipeline(
        csv_content(
            *(
                make_row(
                    sku=f"SKU-{position}",
                    impressions=value,
                    clicks=0,
                    orders=0,
                    units_sold=0,
                    sales=0,
                    ad_spend=0,
                    refunds=0,
                )
                for position, value in enumerate(boundary_values)
            )
        ),
        filename="large-count.csv",
        group_by="sku",
    )
    report_data, _, workbook = workbook_for(result)
    worksheet = workbook["Metrics"]
    columns = column_positions(worksheet)
    rows = {
        worksheet.cell(row=row, column=columns["sku"]).value: row
        for row in range(2, worksheet.max_row + 1)
    }

    for position, value in enumerate(boundary_values):
        sku = f"SKU-{position}"
        metrics_value = report_data.metrics.loc[  # type: ignore[union-attr]
            report_data.metrics["sku"] == sku, "impressions"  # type: ignore[union-attr]
        ].iloc[0]
        cell = worksheet.cell(rows[sku], columns["impressions"])
        assert metrics_value == value
        if abs(value) <= 2**53 - 1:
            assert cell.value == value
            assert cell.data_type == "n"
        else:
            assert cell.value == str(value)
            assert cell.data_type == "s"


def test_large_integer_fallback_applies_to_non_count_columns() -> None:
    report_data = build_report_data(sample_result())
    exact_integer = 2**53
    report_data.metrics["external_reference"] = exact_integer  # type: ignore[index]
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Metrics"]
    cell = worksheet.cell(2, column_positions(worksheet)["external_reference"])

    assert cell.value == str(exact_integer)
    assert cell.data_type == "s"


@pytest.mark.parametrize(
    ("code", "expected_format"),
    [
        (HIGH_IMPRESSIONS_LOW_CTR, PERCENTAGE_FORMAT),
        (LOW_CVR, PERCENTAGE_FORMAT),
        (CLICKS_WITHOUT_ORDERS, COUNT_FORMAT),
        (SPEND_WITHOUT_ORDERS, COUNT_FORMAT),
        (LOW_ROAS, ROAS_FORMAT),
        (HIGH_REFUND_RATE, PERCENTAGE_FORMAT),
        (OUT_OF_STOCK, COUNT_FORMAT),
    ],
)
def test_diagnostics_are_metric_formatted_and_evidence_is_stable_json(
    code: str,
    expected_format: str,
) -> None:
    report_data, _, workbook = workbook_for(sample_result())
    worksheet = workbook["Diagnostics"]
    columns = column_positions(worksheet)
    excel_row = next(
        row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=columns["code"]).value == code
    )
    diagnostic_row = report_data.diagnostics.loc[  # type: ignore[union-attr]
        report_data.diagnostics["code"] == code  # type: ignore[union-attr]
    ].iloc[0]

    for column in ("actual_value", "threshold"):
        cell = worksheet.cell(excel_row, columns[column])
        assert cell.value == pytest.approx(diagnostic_row[column])
        assert cell.number_format == expected_format
    assert worksheet.cell(excel_row, columns["evidence"]).value == json.dumps(
        diagnostic_row["evidence"],
        ensure_ascii=False,
        sort_keys=True,
    )
    assert worksheet.cell(excel_row, columns["message"]).value == diagnostic_row[
        "message"
    ]


def test_evidence_special_scalars_are_serialized_without_mutation() -> None:
    report_data = build_report_data(sample_result())
    evidence = {
        "when": date(1600, 1, 1),
        "count": pd.Series([5], dtype="Int64").iloc[0],
        "missing": pd.NA,
        "nan": float("nan"),
    }
    report_data.diagnostics.at[0, "evidence"] = evidence  # type: ignore[union-attr]
    evidence_before = copy.deepcopy(evidence)

    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Diagnostics"]
    evidence_column = column_positions(worksheet)["evidence"]

    assert worksheet.cell(2, evidence_column).value == (
        '{"count": 5, "missing": null, "nan": null, "when": "1600-01-01"}'
    )
    assert evidence["when"] == evidence_before["when"]
    assert evidence["count"] == evidence_before["count"]
    assert evidence["missing"] is pd.NA
    assert pd.isna(evidence["nan"])


def test_validation_failed_still_exports_four_sheets_with_reason() -> None:
    columns = tuple(column for column in REQUIRED_COLUMNS if column != "sku")
    result = run_pipeline(
        csv_content(make_row(), columns=columns),
        filename="missing-sku.csv",
        group_by="sku",
    )
    report_data, _, workbook = workbook_for(result)
    values = summary_values(report_data)

    assert result.status is PipelineStatus.VALIDATION_FAILED
    assert values[("Pipeline", "Status")] == "VALIDATION_FAILED"
    assert values[("Validation", "Fatal Issues")] == 1
    assert report_data.validation_issues["code"].tolist() == [
        "MISSING_REQUIRED_COLUMN"
    ]
    assert workbook.sheetnames == list(SHEET_NAMES)
    assert workbook["Metrics"]["A1"].value == VALIDATION_FAILED_MESSAGE
    assert workbook["Diagnostics"]["A1"].value == VALIDATION_FAILED_MESSAGE


def test_all_rows_excluded_success_exports_empty_tables_not_failure_message() -> None:
    result = run_pipeline(
        csv_content(make_row(impressions=10, clicks=11)),
        filename="all-excluded.csv",
        group_by="sku",
    )
    report_data, _, workbook = workbook_for(result)

    assert result.status is PipelineStatus.SUCCESS
    assert summary_values(report_data)[("Validation", "Valid Rows")] == 0
    assert report_data.metrics is not None and report_data.metrics.empty
    assert report_data.diagnostics is not None and report_data.diagnostics.empty
    assert [cell.value for cell in workbook["Metrics"][1]] == [
        "sku",
        *BASE_MEASURES,
        *DERIVED_METRICS,
    ]
    assert workbook["Metrics"].max_row == 1
    assert [cell.value for cell in workbook["Diagnostics"][1]] == [
        "sku",
        *DIAGNOSTIC_ISSUE_COLUMNS,
    ]
    assert workbook["Diagnostics"].max_row == 1


def test_no_issues_sheets_keep_headers_without_fake_rows() -> None:
    result = run_pipeline(
        csv_content(
            make_row(
                impressions=100,
                clicks=10,
                orders=1,
                sales=20,
                ad_spend=5,
                refunds=0,
                inventory=10,
            )
        ),
        filename="normal.csv",
        group_by="sku",
    )
    report_data, _, workbook = workbook_for(result)

    assert report_data.validation_issues.empty
    assert report_data.diagnostics is not None and report_data.diagnostics.empty
    assert [cell.value for cell in workbook["Validation Issues"][1]] == list(
        VALIDATION_ISSUE_COLUMNS
    )
    assert workbook["Validation Issues"].max_row == 1
    assert workbook["Diagnostics"].max_row == 1


def test_all_excel_safe_dates_use_native_excel_dates() -> None:
    result = run_pipeline(
        csv_content(
            make_row(date="2025-01-01", sku="SKU-EARLY"),
            make_row(date="2026-08-27", sku="SKU-LATE"),
        ),
        filename="safe-dates.csv",
        group_by="date",
    )
    dates_before = result.metrics["date"].tolist()  # type: ignore[index]
    _, _, workbook = workbook_for(result)
    worksheet = workbook["Metrics"]
    date_column = column_positions(worksheet)["date"]

    assert [worksheet.cell(row, date_column).value for row in (2, 3)] == [
        datetime(2025, 1, 1),
        datetime(2026, 8, 27),
    ]
    assert all(
        worksheet.cell(row, date_column).data_type == "d" for row in (2, 3)
    )
    assert all(
        worksheet.cell(row, date_column).number_format == DATE_FORMAT
        for row in (2, 3)
    )
    assert result.metrics["date"].tolist() == dates_before  # type: ignore[index]
    assert all(type(value) is date for value in dates_before)


def test_wide_date_forces_entire_date_column_to_iso_text() -> None:
    result = run_pipeline(
        csv_content(
            make_row(date="1600-01-01", sku="SKU-EARLY"),
            make_row(date="2026-08-27", sku="SKU-MIDDLE"),
            make_row(date="9999-12-31", sku="SKU-LATE"),
        ),
        filename="wide-dates.csv",
        group_by="date",
    )
    dates_before = result.metrics["date"].tolist()  # type: ignore[index]
    _, _, workbook = workbook_for(result)
    worksheet = workbook["Metrics"]
    date_column = column_positions(worksheet)["date"]

    assert [worksheet.cell(row, date_column).value for row in (2, 3, 4)] == [
        "1600-01-01",
        "2026-08-27",
        "9999-12-31",
    ]
    assert all(
        worksheet.cell(row, date_column).data_type == "s" for row in (2, 3, 4)
    )
    assert all(
        worksheet.cell(row, date_column).number_format == "General"
        for row in (2, 3, 4)
    )
    assert result.metrics["date"].tolist() == dates_before  # type: ignore[index]
    assert all(type(value) is date for value in dates_before)


def test_message_at_excel_text_limit_is_preserved_exactly() -> None:
    report_data = build_report_data(sample_result())
    message = "m" * EXCEL_MAX_CELL_TEXT_LENGTH
    report_data.validation_issues.loc[0, "message"] = message
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Validation Issues"]
    cell = worksheet.cell(2, column_positions(worksheet)["message"])

    assert cell.value == message
    assert len(cell.value) == EXCEL_MAX_CELL_TEXT_LENGTH
    assert cell.data_type == "s"


def test_message_above_excel_text_limit_fails_before_save() -> None:
    report_data = build_report_data(sample_result())
    report_data.validation_issues.loc[0, "message"] = (
        "m" * (EXCEL_MAX_CELL_TEXT_LENGTH + 1)
    )

    with pytest.raises(ReportError) as caught:
        generate_excel_report(report_data)

    assert caught.value.code == EXCEL_CELL_TEXT_TOO_LONG
    assert "Validation Issues" in caught.value.message
    assert "row=2" in caught.value.message
    assert "column=E" in caught.value.message
    assert str(EXCEL_MAX_CELL_TEXT_LENGTH + 1) in caught.value.message
    assert str(EXCEL_MAX_CELL_TEXT_LENGTH) in caught.value.message


def test_serialized_evidence_at_excel_text_limit_is_preserved_exactly() -> None:
    report_data = build_report_data(sample_result())
    evidence = evidence_with_serialized_length(EXCEL_MAX_CELL_TEXT_LENGTH)
    expected = json.dumps(
        evidence,
        ensure_ascii=False,
        sort_keys=True,
        allow_nan=False,
    )
    report_data.diagnostics.at[0, "evidence"] = evidence  # type: ignore[union-attr]
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Diagnostics"]
    cell = worksheet.cell(2, column_positions(worksheet)["evidence"])

    assert len(expected) == EXCEL_MAX_CELL_TEXT_LENGTH
    assert cell.value == expected
    assert cell.data_type == "s"


def test_serialized_evidence_above_excel_text_limit_fails_before_save() -> None:
    report_data = build_report_data(sample_result())
    evidence = evidence_with_serialized_length(EXCEL_MAX_CELL_TEXT_LENGTH + 1)
    report_data.diagnostics.at[0, "evidence"] = evidence  # type: ignore[union-attr]

    with pytest.raises(ReportError) as caught:
        generate_excel_report(report_data)

    assert caught.value.code == EXCEL_CELL_TEXT_TOO_LONG
    assert "Diagnostics" in caught.value.message
    assert "row=2" in caught.value.message
    assert "column=" in caught.value.message
    assert str(EXCEL_MAX_CELL_TEXT_LENGTH + 1) in caught.value.message


def test_evidence_at_maximum_nesting_depth_exports_successfully() -> None:
    report_data = build_report_data(sample_result())
    report_data.diagnostics.at[0, "evidence"] = nested_evidence(  # type: ignore[union-attr]
        MAX_EVIDENCE_DEPTH
    )

    excel_bytes = generate_excel_report(report_data)

    assert excel_bytes.startswith(b"PK")


def test_evidence_above_maximum_nesting_depth_is_rejected() -> None:
    report_data = build_report_data(sample_result())
    report_data.diagnostics.at[0, "evidence"] = nested_evidence(  # type: ignore[union-attr]
        MAX_EVIDENCE_DEPTH + 1
    )

    with pytest.raises(ReportError) as caught:
        generate_excel_report(report_data)

    assert caught.value.code == INVALID_REPORT_DATA
    assert str(MAX_EVIDENCE_DEPTH) in caught.value.message


def test_cyclic_evidence_is_rejected_without_recursion_error() -> None:
    report_data = build_report_data(sample_result())
    evidence: dict[str, object] = {}
    evidence["self"] = evidence
    report_data.diagnostics.at[0, "evidence"] = evidence  # type: ignore[union-attr]

    with pytest.raises(ReportError) as caught:
        generate_excel_report(report_data)

    assert caught.value.code == INVALID_REPORT_DATA
    assert "循环引用" in caught.value.message
    assert caught.value.__cause__ is None


@pytest.mark.parametrize(
    "detail_name",
    ["validation_issues", "metrics", "diagnostics"],
)
def test_excel_row_limit_is_checked_before_workbook_cells_are_created(
    detail_name: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = build_report_data(sample_result())
    oversized = pd.DataFrame(index=pd.RangeIndex(EXCEL_MAX_DATA_ROWS + 1))
    report_data = ReportData(
        summary=source.summary,
        validation_issues=(
            oversized if detail_name == "validation_issues" else source.validation_issues
        ),
        metrics=oversized if detail_name == "metrics" else source.metrics,
        diagnostics=oversized if detail_name == "diagnostics" else source.diagnostics,
    )
    build_called = False

    def fail_if_called(_: ReportData) -> None:
        nonlocal build_called
        build_called = True
        raise AssertionError("workbook construction must not start")

    monkeypatch.setattr(report_module, "_build_workbook", fail_if_called)

    with pytest.raises(ReportError) as caught:
        generate_excel_report(report_data)

    assert caught.value.code == EXCEL_ROW_LIMIT_EXCEEDED
    assert str(EXCEL_MAX_DATA_ROWS + 1) in caught.value.message
    assert build_called is False


def test_excel_row_limit_helper_accepts_maximum_data_rows() -> None:
    maximum = pd.DataFrame(index=pd.RangeIndex(EXCEL_MAX_DATA_ROWS))

    _validate_excel_row_limit(maximum, sheet_name="Metrics")


@pytest.mark.parametrize(
    "mutation",
    [
        "remove_validation_issue",
        "remove_metric_group",
        "remove_diagnostic_issue",
        "change_validation_code",
        "change_diagnostic_code",
        "reorder_validation_code_summary",
        "invalidate_validation_level",
    ],
)
def test_export_rejects_summary_detail_inconsistency(mutation: str) -> None:
    report_data = build_report_data(sample_result())
    if mutation == "remove_validation_issue":
        report_data.validation_issues.drop(index=0, inplace=True)
    elif mutation == "remove_metric_group":
        report_data.metrics.drop(index=0, inplace=True)  # type: ignore[union-attr]
    elif mutation == "remove_diagnostic_issue":
        report_data.diagnostics.drop(index=0, inplace=True)  # type: ignore[union-attr]
    elif mutation == "change_validation_code":
        report_data.validation_issues.loc[0, "code"] = "CHANGED_CODE"
    elif mutation == "change_diagnostic_code":
        report_data.diagnostics.loc[0, "code"] = "CHANGED_CODE"  # type: ignore[union-attr]
    elif mutation == "invalidate_validation_level":
        report_data.validation_issues.loc[0, "level"] = pd.NA
    else:
        code_rows = report_data.summary.index[
            report_data.summary["Section"] == "Validation Code Counts"
        ].tolist()
        first, second = code_rows[:2]
        left = report_data.summary.loc[first].copy()
        report_data.summary.loc[first] = report_data.summary.loc[second]
        report_data.summary.loc[second] = left

    with pytest.raises(ReportError) as caught:
        generate_excel_report(report_data)

    assert caught.value.code == INCONSISTENT_REPORT_DATA


def test_consistency_check_does_not_recalculate_metric_values() -> None:
    report_data = build_report_data(sample_result())
    report_data.metrics.loc[0, "gmv"] = -123.45  # type: ignore[union-attr]
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Metrics"]

    assert worksheet.cell(2, column_positions(worksheet)["gmv"]).value == -123.45


def test_consistency_check_does_not_rederive_warning_rows() -> None:
    report_data = build_report_data(sample_result())
    warning_row = (
        (report_data.summary["Section"] == "Validation")
        & (report_data.summary["Item"] == "Warning Rows")
    )
    report_data.summary.loc[warning_row, "Value"] = 999
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Summary"]
    values = {
        (worksheet.cell(row, 1).value, worksheet.cell(row, 2).value): worksheet.cell(
            row, 3
        ).value
        for row in range(2, worksheet.max_row + 1)
    }

    assert values[("Validation", "Warning Rows")] == 999


def test_object_columns_write_all_missing_sentinels_as_blank() -> None:
    report_data = build_report_data(sample_result())
    values = [pd.NA, float("nan"), None] + ["present"] * (
        len(report_data.metrics) - 3  # type: ignore[arg-type]
    )
    report_data.metrics["optional_note"] = pd.Series(  # type: ignore[index]
        values,
        dtype="object",
    )
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Metrics"]
    column = column_positions(worksheet)["optional_note"]

    assert [worksheet.cell(row, column).value for row in (2, 3, 4)] == [
        None,
        None,
        None,
    ]
    assert worksheet.cell(5, column).value == "present"


def test_validation_none_row_and_field_are_written_as_blank() -> None:
    report_data = build_report_data(sample_result())
    report_data.validation_issues.loc[0, "row"] = pd.NA
    report_data.validation_issues.loc[0, "field"] = None
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Validation Issues"]
    columns = column_positions(worksheet)

    assert worksheet.cell(2, columns["row"]).value is None
    assert worksheet.cell(2, columns["field"]).value is None


def test_illegal_xml_character_is_wrapped_as_excel_export_error() -> None:
    report_data = build_report_data(sample_result())
    report_data.validation_issues.loc[0, "message"] = "bad\x00message"

    with pytest.raises(ReportError) as caught:
        generate_excel_report(report_data)

    assert caught.value.code == EXCEL_EXPORT_ERROR
    assert isinstance(caught.value.__cause__, IllegalCharacterError)


def test_workbook_semantics_are_deterministic_without_comparing_zip_bytes() -> None:
    report_data = build_report_data(sample_result())
    workbooks = [
        load_workbook(BytesIO(generate_excel_report(report_data))) for _ in range(2)
    ]

    assert workbooks[0].sheetnames == workbooks[1].sheetnames
    for sheet_name in SHEET_NAMES:
        left = workbooks[0][sheet_name]
        right = workbooks[1][sheet_name]
        assert [
            [(cell.value, cell.number_format) for cell in row]
            for row in left.iter_rows()
        ] == [
            [(cell.value, cell.number_format) for cell in row]
            for row in right.iter_rows()
        ]


@pytest.mark.parametrize("invalid", [None, {}, pd.DataFrame()])
def test_generate_excel_report_rejects_non_report_data(invalid: object) -> None:
    with pytest.raises(ReportError) as caught:
        generate_excel_report(invalid)  # type: ignore[arg-type]

    assert caught.value.code == INVALID_REPORT_DATA


def test_generate_excel_report_rejects_invalid_report_data_members() -> None:
    invalid = ReportData(
        summary=pd.DataFrame(),
        validation_issues=pd.DataFrame(),
        metrics="not-a-dataframe",  # type: ignore[arg-type]
        diagnostics=None,
    )

    with pytest.raises(ReportError) as caught:
        generate_excel_report(invalid)

    assert caught.value.code == INVALID_REPORT_DATA


def test_excel_export_wraps_ordinary_exception_with_cause(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail(_: ReportData) -> None:
        raise ValueError("writer failed")

    monkeypatch.setattr(report_module, "_build_workbook", fail)

    with pytest.raises(ReportError) as caught:
        generate_excel_report(build_report_data(sample_result()))

    assert caught.value.code == EXCEL_EXPORT_ERROR
    assert isinstance(caught.value.__cause__, ValueError)


def test_existing_report_error_is_not_wrapped(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    expected = ReportError("TEST_REPORT_ERROR", "expected")

    def fail(_: ReportData) -> None:
        raise expected

    monkeypatch.setattr(report_module, "_build_workbook", fail)

    with pytest.raises(ReportError) as caught:
        generate_excel_report(build_report_data(sample_result()))

    assert caught.value is expected


def test_long_text_width_is_capped_and_wrapped() -> None:
    report_data = build_report_data(sample_result())
    report_data.validation_issues.loc[0, "message"] = "长" * 1000
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Validation Issues"]
    message_column = column_positions(worksheet)["message"]
    message_letter = worksheet.cell(1, message_column).column_letter

    assert worksheet.column_dimensions[message_letter].width <= 50
    assert worksheet.cell(2, message_column).alignment.wrap_text is True


@pytest.mark.parametrize(
    "formula_like_value",
    ["=HYPERLINK(...)", "=1+1", "+1+1", "-1+1", "@SUM(1,1)"],
)
def test_group_dimension_text_is_written_as_text_not_excel_formula(
    formula_like_value: str,
) -> None:
    result = run_pipeline(
        csv_content(make_row(sku=formula_like_value)),
        filename="formula-like.csv",
        group_by="sku",
    )
    _, _, workbook = workbook_for(result)
    worksheet = workbook["Metrics"]
    sku_cell = worksheet.cell(2, column_positions(worksheet)["sku"])

    assert sku_cell.value == formula_like_value
    assert sku_cell.data_type == "s"


@pytest.mark.parametrize(
    "formula_like_value",
    ["=HYPERLINK(...)", "=1+1", "+1+1", "-1+1", "@SUM(1,1)"],
)
def test_validation_message_is_written_as_text_not_excel_formula(
    formula_like_value: str,
) -> None:
    report_data = build_report_data(sample_result())
    report_data.validation_issues.loc[0, "message"] = formula_like_value
    workbook = load_workbook(BytesIO(generate_excel_report(report_data)))
    worksheet = workbook["Validation Issues"]
    cell = worksheet.cell(2, column_positions(worksheet)["message"])

    assert cell.value == formula_like_value
    assert cell.data_type == "s"
